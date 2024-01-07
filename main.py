from fastapi import FastAPI, HTTPException,UploadFile, File, Form
from fastapi.responses import JSONResponse
from database import connect_to_database
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import uvicorn
import openpyxl
from fastapi.encoders import jsonable_encoder
from eventCrud import app as event_app
from getChart import chart_app

app = FastAPI()




connection = connect_to_database()
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # You can replace "*" with specific origins if needed
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.include_router(event_app, prefix="/events", tags=["events"])
# app.include_router(attendance_app, prefix="/attendance", tags=["attendance"])
app.include_router(chart_app, prefix="/charts", tags=["charts"])


def table_exists(cursor, table_name):
    cursor.execute(f"SHOW TABLES LIKE '{table_name}'")
    return cursor.fetchone() is not None

def extract_sex_data(df, column_name):
    if column_name not in df.columns:
        raise ValueError(f"Column '{column_name}' not found in the dataframe.")

def load_dataset(file_path):
    file_extension = file_path.split('.')[-1]
    if file_extension == 'csv':
        return pd.read_csv(file_path)
    elif file_extension in ['xls', 'xlsx']:
        return pd.read_excel(file_path)
    else:
        raise ValueError("Unsupported file format. Please provide a CSV or Excel file.")

def read_xlsx(file: UploadFile):
    try:
        workbook = openpyxl.load_workbook(file.file)
        sheet = workbook.active
        sex_column_data = []
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        sex_column_index = None

        for i, value in enumerate(header_row):
            if value == "Sex":
                sex_column_index = i + 1
                break
        if sex_column_index is not None:
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
                sex_column_data.append(row[sex_column_index - 1])
            df = pd.DataFrame({"Sex": sex_column_data})
            sex_counts = df["Sex"].value_counts().to_dict()

            return sex_counts
        else:
            raise HTTPException(status_code=400, detail="Column 'Sex' not found in the XLSX file.")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error reading XLSX file: {str(e)}")


@app.post("/uploadfile/")
async def create_upload_file(file: UploadFile = File(...), title: str = Form(...), date: str = Form(...)):
    if file.filename.endswith(".xlsx"):
        column_data = read_xlsx(file)
        connection = connect_to_database()
        for category, count in column_data.items():
            insert_attendance(category, count, date, title, connection)
        connection.close()

        response_data = {
            "file": jsonable_encoder(column_data),
            "title": title,
            "date": date
        }
        return JSONResponse(content=response_data, media_type="application/json")
    else:
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload an XLSX file.")
    
def insert_attendance(category, count, date_filed, event_name, connection):
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                INSERT INTO attendance (category, count, date_filed, event_name)
                VALUES (%s, %s, %s, %s)
            """, (category, count, date_filed, event_name))  
        connection.commit()  # Commit changes to the database
        # print("Insert successful")
    except Exception as e:
        connection.rollback()  # Rollback changes in case of an error
        print(f"Error inserting data into MySQL: {e}")

if __name__ == "__main__":
    uvicorn.run(app, host="127.0.0.1", port=8000)